# pump_profile.py
from __future__ import annotations

from typing import List, Optional

from scheme.pump_profile import PumpProfile
import os

from openpyxl import load_workbook



def _as_float(x) -> Optional[float]:
    try:
        if x is None:
            return None
        return float(x)
    except Exception:
        return None


def load_pump_profile_xlsx(path: str, sheet_name: str | None = None) -> PumpProfile:
    """
    Columns:
      1) duration
      2) time
      3) rpm
    If 'time' empty -> cumulative sum of duration.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return PumpProfile([], [])

    # header detect
    start_i = 0
    if any(isinstance(v, str) for v in (rows[0][0:3] if len(rows[0]) >= 3 else rows[0])):
        start_i = 1

    t: List[float] = []
    rpm: List[float] = []
    cum_t = 0.0

    for r in rows[start_i:]:
        if not r:
            continue
        dur = _as_float(r[0]) if len(r) > 0 else None
        tt = _as_float(r[1]) if len(r) > 1 else None
        rr = _as_float(r[2]) if len(r) > 2 else None
        if rr is None:
            continue

        if tt is not None:
            cum_t = tt
        else:
            cum_t += (dur if dur is not None else 0.0)

        t.append(float(cum_t))
        rpm.append(float(rr))

    if t and t[0] != 0.0:
        t0 = t[0]
        t = [x - t0 for x in t]

    return PumpProfile(t, rpm)


def interp_profile(profile: PumpProfile, time_s: float) -> float:
    if not profile.t:
        return 0.0
    x = float(time_s)
    if x <= profile.t[0]:
        return profile.rpm[0]
    if x >= profile.t[-1]:
        return profile.rpm[-1]

    for i in range(1, len(profile.t)):
        if x <= profile.t[i]:
            t0, t1 = profile.t[i - 1], profile.t[i]
            y0, y1 = profile.rpm[i - 1], profile.rpm[i]
            if t1 <= t0:
                return y1
            a = (x - t0) / (t1 - t0)
            return y0 + a * (y1 - y0)
    return profile.rpm[-1]
